#!/usr/bin/env python3
"""
whonet_to_template.py

Convert a WHONET-style wide CSV/XLSX into the app's narrow template CSV that
your backend expects (one row per antibiotic result).

Usage:
  # CSV input
  python3 whonet_to_template.py /path/to/whonet_file.csv

  # XLSX input
  python3 whonet_to_template.py /path/to/whonet_file.xlsx

Outputs (same folder as input):
  - whonet_converted.csv
  - whonet_converted.xlsx   (if openpyxl is available)
"""
import csv, sys, io
from pathlib import Path

# Try to import openpyxl for reading/writing XLSX
try:
    from openpyxl import load_workbook, Workbook
    HAVE_OPENPYXL = True
except Exception:
    HAVE_OPENPYXL = False

# canonical expected headers
EXPECTED = {
    "patient_id", "age", "sex", "specimen_type", "organism",
    "antibiotic", "ast_result", "test_date", "host_type",
    "facility", "patient_type", "animal_species", "environment_type"
}

# mapping of common WHONET-ish headers -> expected names
HEADER_MAP = {
    "patient id": "patient_id",
    "patient": "patient_id",
    "lab no": None,             # ignore
    "specimen date": "test_date",
    "date": "test_date",
    "specimen": "specimen_type",
    "sample": "specimen_type",
    "organism": "organism",
    "sex": "sex",
    "age": "age",
    "host type": "host_type",
    "host": "host_type",
    "facility": "facility",
    "patient type": "patient_type",
    "animal species": "animal_species",
    "environment type": "environment_type",
    "notes": "notes"
}

def normalize_header(h):
    if h is None:
        return ""
    return str(h).strip()

def read_rows_from_csv(path):
    with path.open(newline='', encoding='utf-8') as f:
        rdr = csv.DictReader(f)
        headers = [normalize_header(h) for h in rdr.fieldnames or []]
        rows = list(rdr)
    return headers, rows

def read_rows_from_xlsx(path):
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    try:
        header = [normalize_header(x) for x in next(it)]
    except StopIteration:
        return [], []
    rows = []
    for r in it:
        row = {}
        for k, v in zip(header, r):
            row[k] = "" if v is None else str(v)
        rows.append(row)
    return header, rows

def map_headers(headers):
    mapped = []
    for h in headers:
        hu = h.strip().lower()
        if hu in HEADER_MAP:
            tgt = HEADER_MAP[hu]
            if tgt is None:
                mapped.append(None)
            else:
                mapped.append(tgt)
        else:
            # if header already matches expected key, keep it
            if hu in EXPECTED:
                mapped.append(hu)
            else:
                # treat as antibiotic column (uppercase name preserved for output)
                mapped.append(h)
    return mapped

def pivot_to_narrow(headers, mapped_headers, rows):
    """
    headers: original header strings
    mapped_headers: for each header the canonical target (or antibiotic name)
    rows: list of dicts keyed by original header strings
    """
    out_rows = []
    for i, r in enumerate(rows, start=1):
        # basic fields: find values for canonical names (case-sensitive as keys)
        def val_for(name):
            # try exact header match (case-insensitive)
            for h in headers:
                if (h or "").strip().lower() == name:
                    return (r.get(h) or "").strip()
            # else try direct key
            return (r.get(name) or "").strip()

        patient_id = val_for("patient_id") or val_for("patient id") or ""
        age = val_for("age")
        sex = val_for("sex") or "U"
        specimen_type = val_for("specimen_type") or val_for("specimen") or ""
        organism = val_for("organism") or ""
        test_date = val_for("test_date") or val_for("specimen date") or ""
        host_type = val_for("host_type") or val_for("host type") or ""
        facility = val_for("facility") or ""
        patient_type = val_for("patient_type") or val_for("patient type") or ""
        animal_species = val_for("animal_species") or val_for("animal species") or ""
        environment_type = val_for("environment_type") or val_for("environment type") or ""

        # ensure age not empty
        if not age:
            age = "0"

        # for each header that is treated as an antibiotic column, create rows
        for h, mapped in zip(headers, mapped_headers):
            if not mapped or mapped in EXPECTED:
                # not an antibiotic column
                continue
            # mapped is the antibiotic column name (original header preserved)
            v = (r.get(h) or "").strip()
            if v == "":
                continue
            # standardize S/I/R to single letter if possible
            vv = v.strip().upper()
            if vv.startswith("SUS") or vv == "SUSCEPTIBLE": vv = "S"
            if vv.startswith("INT") or vv == "I": vv = "I"
            if vv.startswith("RES") or vv == "R": vv = "R"
            # Accept S/I/R or words; write whatever is present but prefer single char
            out = {
                "patient_id": patient_id,
                "age": age,
                "sex": sex,
                "specimen_type": specimen_type,
                "organism": organism,
                "antibiotic": mapped.strip(),
                "ast_result": vv,
                "test_date": test_date,
                "host_type": host_type,
                "facility": facility,
                "patient_type": patient_type,
                "animal_species": animal_species,
                "environment_type": environment_type
            }
            out_rows.append(out)
    return out_rows

def write_csv(out_rows, outpath):
    fieldnames = ["patient_id","age","sex","specimen_type","organism","antibiotic","ast_result","test_date","host_type","facility","patient_type","animal_species","environment_type"]
    with outpath.open("w", newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in out_rows:
            w.writerow(r)

def write_xlsx(out_rows, outpath):
    if not HAVE_OPENPYXL:
        return
    wb = Workbook()
    ws = wb.active
    fieldnames = ["patient_id","age","sex","specimen_type","organism","antibiotic","ast_result","test_date","host_type","facility","patient_type","animal_species","environment_type"]
    ws.append(fieldnames)
    for r in out_rows:
        ws.append([r.get(c,"") for c in fieldnames])
    wb.save(str(outpath))

def main():
    if len(sys.argv) < 2:
        print("Usage: whonet_to_template.py /path/to/file.csv|xlsx")
        sys.exit(1)
    p = Path(sys.argv[1])
    if not p.exists():
        print("File not found:", p)
        sys.exit(1)

    if p.suffix.lower() in (".xlsx", ".xls") and HAVE_OPENPYXL:
        headers, rows = read_rows_from_xlsx(p)
    else:
        headers, rows = read_rows_from_csv(p)

    headers = [h or "" for h in headers]
    mapped = map_headers(headers)

    out_rows = pivot_to_narrow(headers, mapped, rows)

    out_csv = p.parent / "whonet_converted.csv"
    write_csv(out_rows, out_csv)
    print("WROTE:", out_csv, "(rows:", len(out_rows), ")")

    out_xlsx = p.parent / "whonet_converted.xlsx"
    if HAVE_OPENPYXL:
        write_xlsx(out_rows, out_xlsx)
        print("WROTE:", out_xlsx)

if __name__ == "__main__":
    main()
