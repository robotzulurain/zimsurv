from collections import defaultdict
from datetime import date
from django.db.models import Count, Q
from django.db.models.functions import TruncMonth
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from rest_framework import status
from django.http import HttpResponse

from .models import LabResult

class PublicAPIView(APIView):
    permission_classes = [AllowAny]

# -----------------------
# Helpers
# -----------------------
def filter_qs(request):
    qs = LabResult.objects.all()
    q = request.query_params
    if v := q.get("organism"):   qs = qs.filter(organism=v)
    if v := q.get("antibiotic"): qs = qs.filter(antibiotic=v)
    if v := q.get("facility"):   qs = qs.filter(facility=v)
    if v := q.get("host_type"):  qs = qs.filter(host_type=v)
    return qs

def _parse_date_any(s: str):
    from datetime import datetime
    s = (s or "").strip()
    if not s:
        return None
    # ISO first
    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        pass
    # DD/MM/YYYY
    try:
        return datetime.strptime(s, "%d/%m/%Y").date()
    except Exception:
        return None

# -----------------------
# SUMMARY / CHARTS
# -----------------------
class CountsSummaryView(PublicAPIView):
    def get(self, request):
        qs = filter_qs(request)
        return Response({
            "as_of": date.today().isoformat(),
            "total_tests": qs.count(),
            "unique_patients": qs.values("patient_id").distinct().count(),
            "organisms": qs.values("organism").distinct().count(),
            "antibiotics": qs.values("antibiotic").distinct().count(),
        })

class TimeTrendsView(PublicAPIView):
    def get(self, request):
        qs = filter_qs(request).annotate(m=TruncMonth("test_date")) \
                               .values("m") \
                               .annotate(tests=Count("id"),
                                         resistant=Count("id", filter=Q(ast_result="R"))) \
                               .order_by("m")
        series = []
        for r in qs:
            tests = r["tests"] or 0
            pr = round((r["resistant"]/tests)*100,1) if tests else 0.0
            series.append({"month": r["m"].strftime("%Y-%m"), "tests": tests, "percent_resistant": pr})
        return Response({"series": series})

class AntibiogramView(PublicAPIView):
    def get(self, request):
        qs = filter_qs(request)
        cols = sorted(qs.values_list("antibiotic", flat=True).distinct())
        orgs = sorted(qs.values_list("organism", flat=True).distinct())
        rows = []
        for org in orgs:
            s_list, n_list = [], []
            oq = qs.filter(organism=org).values("antibiotic","ast_result") \
                   .annotate(n=Count("id"))
            lut = {}
            for r in oq:
                ab = r["antibiotic"]; ar = r["ast_result"]; n = r["n"]
                lut.setdefault(ab, {"S":0,"I":0,"R":0,"n":0})
                lut[ab][ar] += n; lut[ab]["n"] += n
            for ab in cols:
                cell = lut.get(ab, {"S":0,"I":0,"R":0,"n":0})
                n = cell["n"]; s = cell["S"]
                s_list.append(round((s/n)*100,1) if n else 0.0)
                n_list.append(n)
            rows.append({"organism": org, "S": s_list, "n": n_list})
        return Response({"columns": cols, "rows": rows, "note": "%S; n isolates"})

class SexAgeView(PublicAPIView):
    def get(self, request):
        qs = filter_qs(request)
        csex = qs.values("sex").annotate(count=Count("id"))
        sex = [{"label": r["sex"] or "Unknown", "count": r["count"]} for r in csex]
        bands = [("0-4",0,4),("5-17",5,17),("18-49",18,49),("50-64",50,64),("65+",65,200)]
        out = []
        for label, lo, hi in bands:
            out.append({"band": label, "count": qs.filter(age__gte=lo, age__lte=hi).count()})
        return Response({"sex": sex, "ageBands": out})

class FacilitiesView(PublicAPIView):
    COORDS = {
        "Harare Central Lab": (-17.8292, 31.0522),
        "Bulawayo Vic Lab": (-20.1325, 28.6265),
        "Gweru Gen Lab": (-19.4500, 29.8167),
    }
    def get(self, request):
        qs = filter_qs(request).values("facility") \
                               .annotate(tests=Count("id"),
                                         resistant=Count("id", filter=Q(ast_result="R"))) \
                               .order_by("facility")
        features = []
        for r in qs:
            tests = r["tests"]; res = r["resistant"]
            pct = round((res/tests)*100,1) if tests else 0.0
            lat,lng = self.COORDS.get(r["facility"], (-19.0,29.8))
            features.append({"facility": r["facility"], "lat":lat,"lng":lng,
                             "samples":tests,"resistance_pct":pct})
        return Response({"features": features})

# -----------------------
# OPTIONS
# -----------------------
class OptionsView(PublicAPIView):
    def get(self, request):
        qs = LabResult.objects.all()
        organisms      = sorted(set(qs.values_list("organism", flat=True)))
        antibiotics    = sorted(set(qs.values_list("antibiotic", flat=True)))
        specimen_types = sorted(set(qs.values_list("specimen_type", flat=True)))
        facilities     = sorted(set(qs.values_list("facility", flat=True)))
        host_types     = ["HUMAN","ANIMAL","ENVIRONMENT"]
        animal_species = ["Cattle","Goat","Sheep","Chicken","Dog","Cat","Wildlife"]
        patient_types  = ["Inpatient","Outpatient","Unknown"]
        sex            = ["M","F","Unknown"]
        ast_results    = ["S","I","R"]
        return Response({
            "organisms": [o for o in organisms if o],
            "antibiotics": [a for a in antibiotics if a],
            "specimen_types": [s for s in specimen_types if s],
            "facilities": [f for f in facilities if f],
            "host_types": host_types,
            "animal_species": animal_species,
            "patient_types": patient_types,
            "sex": sex,
            "ast_results": ast_results,
        })

# -----------------------
# ENTRY / UPLOAD
# -----------------------
class ManualEntryView(PublicAPIView):
    def post(self, request):
        d = request.data
        # Basic validations (prototype)
        req = ["patient_id","specimen_type","organism","antibiotic","ast_result","test_date","facility","host_type"]
        missing = [k for k in req if not str(d.get(k, "")).strip()]
        if missing:
            return Response({"ok": False, "error": f"Missing: {', '.join(missing)}"}, status=400)

        if (d.get("host_type") == "ANIMAL") and not str(d.get("animal_species","")).strip():
            return Response({"ok": False, "error": "animal_species required when host_type=ANIMAL"}, status=400)

        # age
        age_in = d.get("age")
        if age_in not in (None, ""):
            try:
                age = int(age_in)
                if age < 0 or age > 120:
                    return Response({"ok": False, "error": "Age must be 0–120"}, status=400)
            except Exception:
                return Response({"ok": False, "error": "Age must be integer"}, status=400)
        else:
            age = None

        dt = _parse_date_any(d.get("test_date"))
        if not dt:
            return Response({"ok": False, "error": "test_date must be YYYY-MM-DD or DD/MM/YYYY"}, status=400)

        r = LabResult.objects.create(
            patient_id=d.get("patient_id"),
            sex=d.get("sex") or None,
            age=age,
            specimen_type=d.get("specimen_type"),
            organism=d.get("organism"),
            antibiotic=d.get("antibiotic"),
            ast_result=d.get("ast_result"),
            test_date=dt,
            facility=d.get("facility"),
            host_type=d.get("host_type"),
            patient_type=d.get("patient_type") or None,
            animal_species=d.get("animal_species") or None,
        )
        return Response({"ok": True, "id": r.id}, status=201)

class UploadView(PublicAPIView):
    def post(self, request):
        import csv
        f = request.FILES.get("file")
        if not f:
            return Response({"ok": False, "error": "No file uploaded (field 'file')."}, status=400)
        name = (f.name or "").lower()
        rows, errors, imported = [], [], 0

        try:
            if name.endswith(".csv"):
                content = f.read()
                try: text = content.decode("utf-8")
                except Exception: text = content.decode("latin-1")
                reader = csv.DictReader(text.splitlines())
                cols = reader.fieldnames or []
                for i, row in enumerate(reader, start=2):
                    rows.append((i, row))
            else:
                from openpyxl import load_workbook
                wb = load_workbook(filename=f, read_only=True)
                ws = wb.active
                head = next(ws.iter_rows(min_row=1, max_row=1))
                cols = ["" if c.value is None else str(c.value).strip() for c in head]
                idx = {h: i for i, h in enumerate(cols)}
                for i, r in enumerate(ws.iter_rows(min_row=2), start=2):
                    row = {h: ("" if idx.get(h) is None else ("" if r[idx[h]].value is None else str(r[idx[h]].value))) for h in cols}
                    rows.append((i, row))
        except Exception as e:
            return Response({"ok": False, "error": f"Failed to read file: {e}"}, status=400)

        required = ["patient_id","specimen_type","organism","antibiotic","ast_result","test_date","facility","host_type"]
        for line, row in rows:
            miss = [k for k in required if not str(row.get(k,"")).strip()]
            if miss:
                errors.append({"line": line, "error": f"Missing: {', '.join(miss)}"}); continue
            if row.get("host_type") == "ANIMAL" and not str(row.get("animal_species","")).strip():
                errors.append({"line": line, "error": "animal_species required when host_type=ANIMAL"}); continue
            # parse/validate
            age = row.get("age")
            if str(age).strip() == "":
                age = None
            else:
                try:
                    age = int(age)
                    if age < 0 or age > 120:
                        errors.append({"line": line, "error": "Age must be 0–120"}); continue
                except Exception:
                    errors.append({"line": line, "error": "Age must be integer"}); continue
            dt = _parse_date_any(row.get("test_date",""))
            if not dt:
                errors.append({"line": line, "error": "test_date must be YYYY-MM-DD or DD/MM/YYYY"}); continue

            LabResult.objects.create(
                patient_id=row.get("patient_id"),
                sex=row.get("sex") or None,
                age=age,
                specimen_type=row.get("specimen_type"),
                organism=row.get("organism"),
                antibiotic=row.get("antibiotic"),
                ast_result=row.get("ast_result"),
                test_date=dt,
                facility=row.get("facility"),
                host_type=row.get("host_type"),
                patient_type=row.get("patient_type") or None,
                animal_species=row.get("animal_species") or None,
            )
            imported += 1

        return Response({"ok": True, "imported": imported, "errors": errors})

# -----------------------
# TEMPLATE CSV (download)
# -----------------------
class TemplateCSVView(PublicAPIView):
    def get(self, request):
        csv_text = (
            "patient_id,sex,age,specimen_type,organism,antibiotic,ast_result,test_date,facility,host_type,patient_type,animal_species\n"
            "P001,F,34,Urine,Escherichia coli,Ciprofloxacin,S,2025-08-15,Harare Central Lab,HUMAN,Outpatient,\n"
            "A001,,2,Stool,Salmonella enterica,Amoxicillin,R,16/08/2025,Bulawayo Vic Lab,ANIMAL,Unknown,Chicken\n"
        )
        resp = HttpResponse(csv_text, content_type="text/csv")
        resp["Content-Disposition"] = 'attachment; filename="amr_template.csv"'
        return resp

# -----------------------
# ALERTS (simple stub for now)
# -----------------------

class AlertsView(PublicAPIView):
    """
    Heuristics:
      - Rare resistance: organism+antibiotic where overall %S >= 80% but we saw >=1 R last 30 days
      - Spike: last complete month tests > 2 * avg of previous 3 months (same facility if provided)
      - Cluster: >=2 isolates of same organism at a facility within last 14 days
    """
    def get(self, request):
        from datetime import date, timedelta
        from django.db.models import Count, Q
        today = date.today()
        last30 = today - timedelta(days=30)
        last14 = today - timedelta(days=14)

        qs = filter_qs(request)

        alerts = []

        # Rare resistance
        overall = qs.values("organism","antibiotic")                    .annotate(n=Count("id"),
                              s=Count("id", filter=Q(ast_result="S")),
                              r=Count("id", filter=Q(ast_result="R")))
        overall_map = {(r["organism"], r["antibiotic"]): r for r in overall}
        recent_r = qs.filter(test_date__gte=last30, ast_result="R")                     .values("organism","antibiotic")                     .annotate(n=Count("id"))
        for rr in recent_r:
            key = (rr["organism"], rr["antibiotic"])
            tot = overall_map.get(key, {"n":0,"s":0})
            n = tot["n"] or 0
            s = tot["s"] or 0
            pctS = round((s/n)*100,1) if n else 0.0
            if n >= 10 and pctS >= 80.0:  # guard small n
                alerts.append({
                    "type": "Rare resistance",
                    "organism": key[0],
                    "antibiotic": key[1],
                    "count_last_30d": rr["n"],
                })

        # Spike month-over-month
        from django.db.models.functions import TruncMonth
        monthly = qs.annotate(m=TruncMonth("test_date"))                    .values("m")                    .annotate(tests=Count("id"))                    .order_by("m")
        months = list(monthly)
        if len(months) >= 4:
            last = months[-1]["tests"]
            prev3 = [m["tests"] for m in months[-4:-1]]
            avg3 = sum(prev3)/3 if prev3 else 0
            if avg3 and last > 2*avg3:
                alerts.append({
                    "type": "Spike",
                    "month": months[-1]["m"].strftime("%Y-%m"),
                    "tests": last,
                    "prev3_avg": round(avg3,1),
                })

        # Clusters in last 14 days, per facility+organism
        cl = qs.filter(test_date__gte=last14)               .values("facility","organism")               .annotate(n=Count("id"))               .order_by("-n")
        for c in cl:
            if (c["n"] or 0) >= 2:
                alerts.append({
                    "type": "Cluster",
                    "facility": c["facility"],
                    "organism": c["organism"],
                    "count_last_14d": c["n"],
                })

        return Response({"alerts": alerts})

class GlassExportView(PublicAPIView):
    def get(self, request):
        qs = filter_qs(request).values(
            "patient_id","sex","age","specimen_type","organism","antibiotic",
            "ast_result","test_date","facility","host_type","patient_type","animal_species"
        )
        lines = ["patient_id,sex,age,specimen_type,organism,antibiotic,ast_result,test_date,facility,host_type,patient_type,animal_species"]
        for r in qs:
            vals = [
                r.get("patient_id") or "",
                r.get("sex") or "",
                "" if r.get("age") is None else str(r.get("age")),
                r.get("specimen_type") or "",
                r.get("organism") or "",
                r.get("antibiotic") or "",
                r.get("ast_result") or "",
                "" if not r.get("test_date") else r["test_date"].isoformat(),
                r.get("facility") or "",
                r.get("host_type") or "",
                r.get("patient_type") or "",
                r.get("animal_species") or "",
            ]
            # Escape commas/quotes if needed (simple CSV-safe join)
            safe = []
            for v in vals:
                v = str(v)
                if any(c in v for c in [',','"','\n']):
                    v = '"' + v.replace('"','""') + '"'
                safe.append(v)
            lines.append(",".join(safe))
        resp = HttpResponse("\n".join(lines) + "\n", content_type="text/csv")
        resp["Content-Disposition"] = 'attachment; filename="glass_export.csv"'
        return resp
